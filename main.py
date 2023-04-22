from typing import Union
import mysql.connector
# import pandas as pd
import openpyxl
from fastapi import FastAPI

app = FastAPI()

mydb = mysql.connector.connect(
  host="localhost",
  user="user1",
  password="Password$123",
  database="catalogue"
)



# def gettestdata():
#     mycursor.execute("SELECT * FROM TEST")
#     myresult = mycursor.fetchall()
#     data=[]
#     for x in myresult:
#         data.append(x)
#     return data

def inputtable():
    mycursor = mydb.cursor()
    dataframe = openpyxl.load_workbook('items.xlsx')
    df = dataframe.active    


    for row in range(1, df.max_row):
        data=[]
        for col in df.iter_cols(1, df.max_column):
            if (col[row].value) == None:
                col[row].value = ''
            data.append(col[row].value)
            
        print("INSERT INTO ITEMS VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)", tuple(data))
        mycursor.execute("INSERT INTO ITEMS VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)", tuple(data))

    mydb.commit()

    mycursor.close()


@app.get("/")
def read_root():
    return {"Hello": "World"}


@app.get("/q0/")
def read_root():
    dataframe = openpyxl.load_workbook('items.xlsx')
    # data = dataframe['source']
    df = dataframe.active 
    rows = [ [cell.value for cell in row if cell.value is not None] for row in df.rows ]
    print(rows)
    return rows

@app.get("/insert/")
def read_root():
    inputtable()
    # print(df)
    # return df


@app.get("/items/")
def read_root():
    mycursor = mydb.cursor()
    mycursor.execute("SELECT * FROM ITEMS")
    data = mycursor.fetchall()
    print(data)
    mycursor.close()
    return data

@app.get("/q1/{item_id}")
def read_item(item_id: str):
    print(item_id)
    mycursor = mydb.cursor()
    mycursor.execute("SELECT parent_code FROM ITEMS WHERE item_code=%s OR item_name=%s ORDER BY parent_code ASC LIMIT 1", (item_id, item_id))
    data = mycursor.fetchall()
    print(data)
    mycursor.close()
    return data

@app.get("/q2/{item_id}")
def read_item(item_id: str):
    print(item_id)
    mycursor = mydb.cursor()
    mycursor.execute("SELECT upc FROM ITEMS WHERE item_name=%s ORDER BY upc ASC ", (item_id,))
    data = mycursor.fetchall()
    print(data)
    mycursor.close()
    return data

@app.get("/q3/")
def read_item():
    mycursor = mydb.cursor()
    mycursor.execute("SELECT count(*) FROM ITEMS WHERE enabled='YES' OR enabled='Y' ")
    active = mycursor.fetchone()
    mycursor.execute("SELECT count(*) FROM ITEMS WHERE enabled='NO' OR enabled='N' ")
    inactive = mycursor.fetchone()
    mycursor.close()
    return {'active': active[0], 'inactive':inactive[0]}


@app.get("/q4/")
def read_item():
    mycursor = mydb.cursor()
    mycursor.execute("SELECT category_l1,AVG(mrp) FROM ITEMS GROUP BY category_l1")
    cat1 = mycursor.fetchall()
    mycursor.execute("SELECT category_l2,AVG(mrp) FROM ITEMS GROUP BY category_l2 ")
    cat2 = mycursor.fetchall()
    mycursor.close()
    return {'category_L1': cat1, 'category_L2':cat2}

# mydb.close()